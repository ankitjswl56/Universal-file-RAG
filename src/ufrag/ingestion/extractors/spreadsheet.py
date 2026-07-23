import re
from dataclasses import dataclass, field

import openpyxl

_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")
_SUM_RANGE = re.compile(r"SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", re.IGNORECASE)
_BINARY_OP = re.compile(r"([A-Z]+\d+)\s*([+\-*/])\s*([A-Z]+\d+)")


@dataclass
class SheetData:
    name: str
    headers: list[str]
    rows: list[list]  # evaluated values, one list per data row (header row excluded)
    formulas: list[tuple[str, str, object]] = field(default_factory=list)  # (cell, formula, value)


def extract_spreadsheet(path) -> list[SheetData]:
    # openpyxl can't give formulas and cached values from a single load, so load twice.
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)

    sheets = []
    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        formula_rows = list(ws_formulas.iter_rows())
        value_rows = list(ws_values.iter_rows())
        if not formula_rows:
            sheets.append(SheetData(name=sheet_name, headers=[], rows=[]))
            continue

        headers = [str(c.value) if c.value is not None else "" for c in formula_rows[0]]

        # coord -> value, seeded with every non-formula cell, used to fall back to a
        # best-effort evaluation when openpyxl has no cached value for a formula cell
        # (only happens when a file was never opened/saved by a real spreadsheet app,
        # e.g. one generated straight from openpyxl or a similar export pipeline).
        known_values: dict[str, object] = {}
        for row_f in formula_rows[1:]:
            for cell in row_f:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    known_values[cell.coordinate] = cell.value

        pending: list[tuple[str, str]] = []  # (coordinate, formula)
        cached: dict[str, object] = {}
        for row_f, row_v in zip(formula_rows[1:], value_rows[1:]):
            for cell_f, cell_v in zip(row_f, row_v):
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    if cell_v.value is not None:
                        cached[cell_f.coordinate] = cell_v.value
                        known_values[cell_f.coordinate] = cell_v.value
                    else:
                        pending.append((cell_f.coordinate, cell_f.value))

        resolved = _resolve_formulas(pending, known_values)

        rows = []
        formulas = []
        for row_f in formula_rows[1:]:
            row_values = []
            for cell_f in row_f:
                coord = cell_f.coordinate
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    value = cached.get(coord, resolved.get(coord))
                    formulas.append((coord, cell_f.value, value))
                    row_values.append(value)
                else:
                    row_values.append(cell_f.value)
            rows.append(row_values)

        sheets.append(SheetData(name=sheet_name, headers=headers, rows=rows, formulas=formulas))
    return sheets


def _resolve_formulas(pending: list[tuple[str, str]], known_values: dict[str, object]) -> dict[str, object]:
    """Best-effort evaluator for exactly two patterns: SUM over a same-column range,

    and a single binary arithmetic op between two cell refs. Anything more complex
    is left unresolved (None) rather than guessed at — this is a fallback for files
    with no cached values at all, not a general formula engine.
    """
    resolved: dict[str, object] = {}
    remaining = list(pending)
    for _ in range(len(pending) + 1):
        if not remaining:
            break
        progressed = False
        still_pending = []
        for coord, formula in remaining:
            value = _try_evaluate(formula, known_values)
            if value is not None:
                resolved[coord] = value
                known_values[coord] = value
                progressed = True
            else:
                still_pending.append((coord, formula))
        remaining = still_pending
        if not progressed:
            break
    return resolved


def _try_evaluate(formula: str, known_values: dict[str, object]) -> object:
    expr = formula.lstrip("=").strip()

    sum_match = _SUM_RANGE.fullmatch(expr)
    if sum_match:
        coords = _expand_column_range(sum_match.group(1), sum_match.group(2))
        if coords and all(known_values.get(c) is not None for c in coords):
            return sum(known_values[c] for c in coords)
        return None

    op_match = _BINARY_OP.fullmatch(expr)
    if op_match:
        left, op, right = op_match.groups()
        a, b = known_values.get(left), known_values.get(right)
        if a is None or b is None:
            return None
        if op == "+":
            return a + b
        if op == "-":
            return a - b
        if op == "*":
            return a * b
        if op == "/":
            return a / b if b else None

    return None


def _expand_column_range(start: str, end: str) -> list[str] | None:
    start_match, end_match = _CELL_REF.match(start), _CELL_REF.match(end)
    if not start_match or not end_match:
        return None
    col_s, row_s = start_match.groups()
    col_e, row_e = end_match.groups()
    if col_s != col_e:
        return None  # only same-column ranges are supported
    return [f"{col_s}{r}" for r in range(int(row_s), int(row_e) + 1)]
